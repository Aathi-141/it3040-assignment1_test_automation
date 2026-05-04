from playwright.sync_api import sync_playwright
import openpyxl
import argparse
import time
import os

def run():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--url", required=True)
    parser.add_argument("--wait-ms", type=int, default=8000)
    parser.add_argument("--type-delay-ms", type=int, default=80)
    parser.add_argument("--slow-mo-ms", type=int, default=200)
    parser.add_argument("--save-every", type=int, default=1)
    parser.add_argument("--keep-open", action="store_true")
    parser.add_argument("--input-col", default="Input")
    parser.add_argument("--expected-col", default="Expected output")
    parser.add_argument("--actual-col", default="Actual output")
    parser.add_argument("--status-col", default="Status")
    args = parser.parse_args()

    # Find excel file
    excel_path = args.excel
    if not os.path.exists(excel_path):
        # try same folder as script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(script_dir, args.excel)
    
    if not os.path.exists(excel_path):
        print(f"ERROR: Cannot find Excel file: {excel_path}")
        return

    print(f"Using Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Find columns by header name
    headers = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}
    print(f"Found headers: {list(headers.keys())}")

    input_col = headers.get(args.input_col)
    expected_col = headers.get(args.expected_col)
    actual_col = headers.get(args.actual_col)
    status_col = headers.get(args.status_col)

    if not input_col:
        print(f"ERROR: Cannot find column '{args.input_col}'")
        return

    print(f"Columns - Input:{input_col} Expected:{expected_col} Actual:{actual_col} Status:{status_col}")
    print(f"Total rows: {ws.max_row - 1}")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            slow_mo=args.slow_mo_ms,
            channel="chrome"  # use installed Chrome
        )
        page = browser.new_page()
        page.set_default_timeout(60000)
        
        print(f"Opening: {args.url}")
        page.goto(args.url, wait_until="domcontentloaded")
        page.wait_for_timeout(3000)
        
        # Find input textarea
        input_box = page.locator('textarea[placeholder*="English"]').first
        input_box.wait_for(state="visible", timeout=30000)
        print("Found input box!")

        for row_num in range(2, ws.max_row + 1):
            input_val = ws.cell(row=row_num, column=input_col).value
            if not input_val:
                continue
            
            singlish = str(input_val).strip()
            expected = ""
            if expected_col:
                ev = ws.cell(row=row_num, column=expected_col).value
                if ev:
                    expected = str(ev).strip()

            print(f"Testing [Row {row_num}]: {singlish[:60]}...")

            try:
                # Clear and type input
                input_box.click()
                page.keyboard.press("Control+A")
                page.keyboard.press("Delete")
                page.wait_for_timeout(500)
                input_box.type(singlish, delay=args.type_delay_ms)
                page.wait_for_timeout(1000)

                # Click Transliterate button
                btn = page.get_by_role("button", name="Transliterate")
                btn.click()
                page.wait_for_timeout(args.wait_ms)

                # Get output - try multiple selectors for the output textarea
                actual = ""
                
                # Try 1: output textarea by placeholder
                try:
                    out = page.locator('textarea[placeholder*="Sinhala"]').first
                    actual = out.evaluate("el => el.value")
                    if actual:
                        actual = actual.strip()
                except:
                    pass

                # Try 2: all textareas, get second one
                if not actual:
                    try:
                        textareas = page.locator("textarea").all()
                        if len(textareas) >= 2:
                            actual = textareas[1].evaluate("el => el.value")
                            if actual:
                                actual = actual.strip()
                    except:
                        pass

                # Try 3: evaluate directly on page
                if not actual:
                    try:
                        actual = page.evaluate("""() => {
                            const textareas = document.querySelectorAll('textarea');
                            for (let i = 0; i < textareas.length; i++) {
                                const ph = textareas[i].getAttribute('placeholder') || '';
                                if (ph.toLowerCase().includes('sinhala') || ph.toLowerCase().includes('transliterat')) {
                                    return textareas[i].value;
                                }
                            }
                            if (textareas.length >= 2) return textareas[1].value;
                            return '';
                        }""")
                        if actual:
                            actual = actual.strip()
                    except:
                        pass

                print(f"  Actual output: '{actual[:50]}'" if actual else "  Actual output: (empty)")

                # Write to Excel
                if actual_col:
                    ws.cell(row=row_num, column=actual_col).value = actual

                if status_col:
                    if actual and expected:
                        status = "PASS" if actual == expected else "FAIL"
                    elif actual:
                        status = "FAIL"  # negative test case, any output = fail
                    else:
                        status = "UI Error"
                    ws.cell(row=row_num, column=status_col).value = status
                    print(f"  -> {status}")

                # Save after every row
                wb.save(excel_path)

            except Exception as e:
                print(f"  ERROR: {e}")
                if status_col:
                    ws.cell(row=row_num, column=status_col).value = "UI Error"
                wb.save(excel_path)

        print("\nAll done! Saving final Excel...")
        wb.save(excel_path)
        print(f"Saved to: {excel_path}")

        if args.keep_open:
            print("Keeping browser open. Press CTRL+C to close.")
            try:
                while True:
                    time.sleep(1)
            except KeyboardInterrupt:
                pass

        browser.close()

if __name__ == "__main__":
    run()
